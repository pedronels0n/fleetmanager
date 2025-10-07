# -------------------- IMPORTS --------------------
from django.template import loader
from django.template.loader import render_to_string
from django.http import HttpResponse, HttpResponseForbidden
from datetime import datetime, timezone, timedelta
from weasyprint import HTML
from .models import Motorista, Veiculo, TermoResponsabilidade, Setor, Multa
from django.core.files.base import ContentFile
from django.shortcuts import render, redirect, get_object_or_404
from .forms import *
import base64
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q  # Para filtros complexos
from .utils import grupo_administrador
from django.utils.text import slugify
from itertools import chain
from django.contrib.auth.models import User 
from openpyxl import Workbook
from openpyxl.styles import Font

# ==================================================
# =================== HOME =========================
# ==================================================

@login_required
def home(request):
    """Página inicial do sistema."""
    return render(request, "controle/index.html")

# ==================================================
# ========== TERMO DE RESPONSABILIDADE =============
# ==================================================

@login_required
def criar_termo(request, pk):
    """
    Cria um termo de responsabilidade para um motorista.
    Salva a assinatura como arquivo.
    """
    motorista = get_object_or_404(Motorista, pk=pk)

    if request.method == "POST":
        form = TermoResponsabilidadeForm(request.POST)
        assinatura_base64 = request.POST.get("assinatura_imagem")

        if not assinatura_base64:
            messages.error(request, "Assinatura não enviada. Por favor, assine antes de salvar.")
            return redirect("assinatura", pk=motorista.id)

        if form.is_valid():
            termo = form.save(commit=False)
            termo.motorista = motorista

            try:
                format, imgstr = assinatura_base64.split(';base64,')
                ext = format.split('/')[-1]
                nome_arquivo = f"assinatura_{slugify(motorista.nome)}.{ext}"
                termo.arquivo.save(nome_arquivo, ContentFile(base64.b64decode(imgstr)), save=False)
            except Exception as e:
                messages.error(request, f"Erro ao processar assinatura: {e}")
                return redirect("assinatura", pk=motorista.id)

            termo.save()

            if termo.veiculo:
                termo.veiculo.motoristas.add(motorista)

            messages.success(request, "Termo salvo com sucesso!")
            return redirect("listar_termos", pk=motorista.id)
    else:
        form = TermoResponsabilidadeForm(initial={'motorista': motorista})

    return render(request, "controle/assinatura.html", {"form": form, "motorista": motorista})

@login_required
def visualizar_termos(request, pk):
    """
    Visualiza todos os termos de responsabilidade de um motorista.
    """
    motorista = get_object_or_404(Motorista, pk=pk)
    termos = TermoResponsabilidade.objects.filter(motorista=motorista)
    dicionario = {"termos" : termos, "motorista" : motorista}
    return render(request, "controle/termos.html", dicionario)

@login_required
def lista_termo(request):
    """
    Lista todos os termos de responsabilidade do sistema.
    """
    termos = TermoResponsabilidade.objects.all()
    dicionario = {"termos" : termos}
    return render(request, "controle/termos.html", dicionario)

@login_required
def termo_responsabilidade(request, veiculo_id, motorista_id):
    """
    Gera o PDF do termo de responsabilidade assinado.
    """
    motorista = Motorista.objects.get(id=motorista_id)
    veiculo = Veiculo.objects.get(id=veiculo_id)
    termo = TermoResponsabilidade.objects.filter(veiculo_id=veiculo_id, motorista_id=motorista_id).first()

    template = loader.get_template("controle/termo_responsabilidade.html")

    # Carrega imagem da assinatura
    with termo.arquivo.open("rb") as f:
        assinatura_base64 = base64.b64encode(f.read()).decode("utf-8")
    img_data = f"data:image/png;base64,{assinatura_base64}"

    html_string = template.render({
        'motorista': motorista,
        'veiculo': veiculo,
        'data_hoje': termo.data_assinatura.strftime("%d/%m/%Y"),
        'assinatura' : img_data 
    })

    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename="termo_{motorista.nome}_{veiculo.placa}.pdf"'
    return response

# ==================================================
# ================= MOTORISTA ======================
# ==================================================

@login_required
def lista_motoristas(request):
    """
    Lista todos os motoristas cadastrados, com ordenação e filtro.
    """
    motoristas = Motorista.objects.all()
    ordem = request.GET.get("ordenar", "a-z")
    if ordem == "z-a":
        motoristas = motoristas.order_by("-nome")
    else:
        motoristas = motoristas.order_by("nome")

    query = request.GET.get("q")
    if query:
        motoristas = motoristas.filter(
            Q(nome__icontains=query) | Q(cpf__icontains=query) | Q(cnh_numero__icontains=query)
        )

    dicionario = {
        "motoristas": motoristas,
        "ordenar": ordem,
        "query": query or "",
    }
    return render(request, "controle/motorista.html", dicionario)

@login_required
def criar_motorista(request):
    """
    Cria um novo motorista.
    """
    if request.method == "POST":
        form = MotoristaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("lista_motorista")
    else:
        form = MotoristaForm()
    return render(request, "controle/motorista_form.html", {"form": form, "titulo": "Criar Condutor"})

@login_required
def editar_motorista(request, pk):
    """
    Edita os dados de um motorista existente.
    """
    motorista = get_object_or_404(Motorista, pk=pk)
    if request.method == "POST":
        form = MotoristaForm(request.POST, request.FILES, instance=motorista)
        if form.is_valid():
            form.save()
            messages.success(request, "Motorista atualizado com sucesso.")
            return redirect("lista_motorista")
    else:
        form = MotoristaForm(instance=motorista)
    return render(request, "controle/motorista_form.html", {"form": form, "titulo": "Editar Motorista", "motorista": motorista})

@user_passes_test(grupo_administrador)
def excluir_motorista(request, pk):
    """
    Exclui um motorista do sistema.
    """
    motorista = get_object_or_404(Motorista, pk=pk)
    if request.method == "POST":
        motorista.delete()
        return redirect("lista_motorista")
    return redirect("lista_motorista")

# ==================================================
# ================= VEICULO ========================
# ==================================================

@login_required
def listar_veiculo(request):
    """
    Lista todos os veículos, com filtros por placa, setor e status.
    """
    if request.user.is_superuser:
        veiculos = Veiculo.objects.all()
    else:
        veiculos = Veiculo.objects.filter(created_by=request.user)

    query = request.GET.get("q")
    if query:
        veiculos = veiculos.filter(placa__icontains=query)

    setor_id = request.GET.get("setor")
    if setor_id:
        veiculos = veiculos.filter(setor_id=setor_id)

    status = request.GET.get("status")
    if status:
        veiculos = veiculos.filter(status_atual=status)

    setores = Setor.objects.all()
    status_choices = Veiculo.STATUS_CHOICE.choices

    contexto = {
        "veiculo": veiculos,
        "query": query or "",
        "setores": setores,
        "setor_id": setor_id or "",
        "status_choices": status_choices,
        "status": status or "",
    }
    return render(request, "controle/veiculo.html", contexto)

@login_required
def detalhar_veiculo(request, pk):
    """
    Mostra detalhes de um veículo e seus últimos abastecimentos.
    """
    veiculo = get_object_or_404(Veiculo, pk=pk)
    abastecimentos = veiculo.abastecimentos.order_by('-data')[:5]  # Últimos 5 abastecimentos
    contexto = {
        "veiculo": veiculo,
        "motoristas": veiculo.motoristas.all(),
        "abastecimentos": abastecimentos,
    }
    return render(request, "controle/detalhar_veiculo.html", contexto)

@login_required
def criar_veiculo(request):
    """
    Cria um novo veículo.
    """
    if request.method == 'POST':
        form = VeiculoForm(request.POST, request.FILES)
        if form.is_valid():
            veiculo = form.save(commit=False)
            veiculo.created_by = request.user
            veiculo.save()
            return redirect('listar_veiculo')
    else:
        form = VeiculoForm()
    return render(request, 'controle/veiculo_form.html', {'form': form, 'titulo': 'Criar Veículo'})

@login_required
def editar_veiculo(request, pk):
    """
    Edita os dados de um veículo existente.
    """
    veiculo = get_object_or_404(Veiculo, pk=pk)
    if request.user != veiculo.created_by and not request.user.is_superuser:
        return HttpResponseForbidden("Você não tem permissão para editar este veículo.")

    if request.method == 'POST':
        form = VeiculoForm(request.POST, request.FILES, instance=veiculo)
        if form.is_valid():
            veiculo = form.save(commit=False)
            veiculo.save()
            form.save_m2m()
            return redirect('listar_veiculo')
    else:
        form = VeiculoForm(instance=veiculo)
    return render(request, 'controle/veiculo_form.html', {'form': form, 'titulo': 'Editar Veículo'})

@login_required
def excluir_veiculo(request, pk):
    """
    Inativa (não exclui fisicamente) um veículo.
    """
    veiculo = get_object_or_404(Veiculo, pk=pk)
    if request.user != veiculo.created_by and not request.user.is_superuser:
        return HttpResponseForbidden("Você não tem permissão para excluir este veículo.")

    if request.method == 'POST':
        veiculo.status_atual = Veiculo.STATUS_CHOICE.INATIVO
        veiculo.save()
        return redirect('listar_veiculo')
    return redirect('listar_veiculo')

# ==================================================
# ================= LOGS ===========================
# ==================================================

@user_passes_test(grupo_administrador, login_url='acesso_negado')
def logs_todos(request):
    """
    Lista todos os logs de alterações dos modelos principais.
    """
    logs_motorista = Motorista.history.all()
    logs_veiculo = Veiculo.history.all()
    logs_termo = TermoResponsabilidade.history.all()
    logs_multas = Multa.history.all()

    todos_logs = sorted(
        chain(logs_motorista, logs_veiculo, logs_termo, logs_multas),
        key=lambda log: log.history_date or timezone.now(),
        reverse=True
    )

    logs_processados = []
    for log in todos_logs:
        modelo = getattr(log.instance._meta, 'model_name', 'desconhecido') if hasattr(log, 'instance') and log.instance else 'desconhecido'
        tipos = {'+': 'Criado', '~': 'Atualizado', '-': 'Deletado'}
        tipo_operacao = tipos.get(log.history_type, 'Desconhecido')

        try:
            if modelo == 'motorista':
                descricao_objeto = getattr(log, 'nome', str(log))
            elif modelo == 'veiculo':
                descricao_objeto = getattr(log, 'placa', str(log))
            elif modelo == 'termoresponsabilidade':
                try:
                    veiculo = log.veiculo
                    motorista = log.motorista
                    descricao_objeto = f"Termo - {veiculo} - {motorista}"
                except Exception:
                    descricao_objeto = "Termo - [dados indisponíveis]"
            else:
                descricao_objeto = str(log)
        except Exception:
            descricao_objeto = 'Erro ao processar descrição'

        logs_processados.append({
            'modelo': modelo.capitalize(),
            'descricao': descricao_objeto,
            'tipo_operacao': tipo_operacao,
            'usuario': getattr(log, 'history_user', None),
            'data': log.history_date,
            'motivo': getattr(log, 'history_change_reason', None),
        })

    return render(request, 'controle/logs_todos.html', {'logs': logs_processados})

# ==================================================
# ============= ABASTECIMENTO ======================
# ==================================================

@user_passes_test(grupo_administrador, login_url='acesso_negado')
def lista_abastecimentos(request):
    """
    Lista todos os abastecimentos, com filtros por placa e data.
    """
    from .models import Abastecimento  # Certifique-se que o model está importado

    hoje = datetime.now().date()
    trinta_dias_atras = hoje - timedelta(days=30)
    abastecimentos = Abastecimento.objects.filter(data__gte=trinta_dias_atras)

    placa = request.GET.get("placa")
    if placa:
        abastecimentos = abastecimentos.filter(veiculo__placa__icontains=placa)

    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    if data_inicio and data_fim:
        try:
            data_inicio_dt = datetime.strptime(data_inicio, "%Y-%m-%d").date()
            data_fim_dt = datetime.strptime(data_fim, "%Y-%m-%d").date()
            abastecimentos = abastecimentos.filter(data__range=(data_inicio_dt, data_fim_dt))
        except ValueError:
            messages.error(request, "Formato de data inválido. Use AAAA-MM-DD.")

    contexto = {
        "abastecimentos": abastecimentos,
        "placa": placa or "",
        "data_inicio": data_inicio or "",
        "data_fim": data_fim or "",
    }
    return render(request, "controle/abastecimento.html", contexto)


def registrar_abastecimento(request):
    """
    Registra um novo abastecimento.
    """
    if request.method == 'POST':
        form = AbastecimentoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('abastecimentos')
    else:
        form = AbastecimentoForm()
    return render(request, 'controle/abastecimento_form.html', {'form': form})

from .models import Abastecimento

@login_required
def editar_abastecimento(request, pk):
    """
    Edita um abastecimento existente.
    """
    abastecimento = get_object_or_404(Abastecimento, pk=pk)
    if request.method == 'POST':
        form = AbastecimentoForm(request.POST, instance=abastecimento)
        if form.is_valid():
            form.save()
            messages.success(request, "Abastecimento atualizado com sucesso!")
            return redirect('abastecimentos')
    else:
        form = AbastecimentoForm(instance=abastecimento)
    return render(request, 'controle/abastecimento_form.html', {'form': form, 'abastecimento': abastecimento})

@login_required
def excluir_abastecimento(request, pk):
    """
    Exclui um abastecimento.
    """
    abastecimento = get_object_or_404(Abastecimento, pk=pk)
    if request.method == 'POST':
        abastecimento.delete()
        messages.success(request, "Abastecimento excluído com sucesso!")
        return redirect('abastecimentos')
    return render(request, 'controle/confirmar_exclusao_abastecimento.html', {'abastecimento': abastecimento})


# ==================================================
# ================ USUÁRIOS ========================
# ==================================================

@user_passes_test(grupo_administrador, login_url='acesso_negado')
def lista_usuarios(request):
    """
    Lista todos os usuários do sistema.
    """
    usuarios = User.objects.all()
    return render(request, "controle/listar_usuarios.html", {"usuarios": usuarios})


@user_passes_test(grupo_administrador, login_url='acesso_negado')
def criar_usuario(request):
    if request.method == "POST":
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuário criado com sucesso!")
            return redirect("lista_usuarios")
        else:
            form = UserForm()
            return render(request, "controle/usuarios_form.html", {"form": form, "titulo": "Criar Usuário"})


@user_passes_test(grupo_administrador, login_url='acesso_negado')
def editar_usuario(request, pk):
    """
    Edita os dados de um usuário existente.
    """
    user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuário atualizado com sucesso!")
            return redirect("lista_usuarios")
    else:
        form = UserForm(instance=user)
    return render(request, "controle/usuarios_form.html", {"form": form, "titulo": "Editar Usuário"})

@user_passes_test(grupo_administrador, login_url='acesso_negado')
def deletar_usuario(request, pk):
    """
    Exclui um usuário do sistema.
    """
    user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        user.delete()
        messages.success(request, "Usuário deletado com sucesso!")
        return redirect("lista_usuarios")
    return render(request, "usuarios/confirm_delete.html", {"usuario": user})

def acesso_negado(request):
    """
    Página de acesso negado.
    """
    return render(request, 'controle/acesso_negado.html', status=403)

# ==================================================
# ================= MULTAS =========================
# ==================================================

@login_required
def criar_multa(request):
    """
    Cria uma nova multa.
    """
    if request.method == 'POST':
        form = MultaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('listar_multas')
    else:
        form = MultaForm()
    return render(request, 'controle/multa_form.html', {'form': form})

@login_required
def atualizar_status_multa(request, pk):
    """
    Atualiza o status de uma multa.
    """
    multa = get_object_or_404(Multa, pk=pk)
    if request.method == "POST":
        form = AtualizarStatusMultaForm(request.POST, request.FILES, instance=multa)
        if form.is_valid():
            form.save()
            messages.success(request, "Status da multa atualizado com sucesso.")
            return redirect("listar_multas")
        else:
            messages.error(request, "Corrija os erros abaixo.")
    else:
        form = AtualizarStatusMultaForm(instance=multa)
    return render(request, "controle/multa_editar.html", {"form": form, "multa": multa})

@login_required
def criar_memorando(request, multa_id):
    """
    Gera o PDF do memorando de uma multa.
    """
    multa = get_object_or_404(Multa, pk=multa_id)
    html_string = render_to_string("controle/memorando.html", {"multa": multa})
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response['Content-Disposition'] = f'inline; filename=memorando_{multa.id}.pdf'
    return response

@login_required
def listar_multas(request):
    """
    Lista todas as multas, com filtros por status e setor.
    """
    multas = Multa.objects.all().order_by('-data_hora_infracao')
    status_multa = request.GET.get('status_multa')
    status_pagamento = request.GET.get('status_pagamento')
    setor = request.GET.get('setor')

    if status_multa and status_multa != "todos":
        multas = multas.filter(status_multa=status_multa)
    if status_pagamento and status_pagamento != "todos":
        multas = multas.filter(status_pagamento=status_pagamento)
    if setor and setor != "todos":
        multas = multas.filter(setor=setor)

    conta_form = ContaPagamentoForm()
    context = {
        'multas': multas,
        'status_multa_selecionado': status_multa,
        'status_pagamento_selecionado': status_pagamento,
        'setor_selecionado': setor,
        'setores': Multa.objects.values_list('setor', flat=True).distinct(),
        'conta_form': conta_form,
    }
    return render(request, 'controle/listar_multas.html', context)

@login_required
def pagar_multa(request, pk):
    """
    Registra o pagamento de uma multa.
    """
    multa = get_object_or_404(Multa, pk=pk)
    if request.method == "POST":
        form = PagamentoMultaForm(request.POST, request.FILES, instance=multa)
        if form.is_valid():
            form.save()
            messages.success(request, "Pagamento registrado com sucesso.")
            return redirect("listar_multas")
        else:
            messages.error(request, "Corrija os erros abaixo.")
    else:
        form = PagamentoMultaForm(instance=multa)
    return render(request, "controle/pagar_multa.html", {"form": form, "multa": multa})

@login_required
def detalhar_multa(request, pk):
    """
    Mostra detalhes de uma multa.
    """
    multa = get_object_or_404(Multa, pk=pk)
    return render(request, "controle/detalhar_multa.html", {"multa": multa})

def trocar_setor(request, pk):
    """
    Troca o setor responsável por uma multa.
    """
    multa = get_object_or_404(Multa, id=pk)
    setores = Setor.objects.all()
    if request.method == 'POST':
        form = TrocaSetorForm(request.POST, instance=multa)
        if form.is_valid():
            form.save()
            return redirect('listar_multas')
    else:
        form = TrocaSetorForm(instance=multa)
    return render(request, 'controle/setor_multa.html', {'form': form, 'multa': multa, 'setores': setores})

# ==================================================
# ================ RELATÓRIOS ======================
# ==================================================

@user_passes_test(grupo_administrador, login_url='acesso_negado')
def exportar_multas_excel(request):
    """
    Exporta todas as multas para um arquivo Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Multas"

    # Cabeçalhos
    ws.append([
        'ID', 'Placa', 'Setor', 'Setor Descrição', 'Motorista',
        'Infração', 'Valor', 'Gravidade',
        'Data/Hora Infração', 'Local', 'Órgão Autuador',
        'Status Multa', 'Status Pagamento', 'Prazo Pagamento', 'Data Registro',
        'Documento Recebido', 'Comprovante Pagamento', 'Notificação Infração'
    ])

    link_font = Font(color="0000FF", underline="single")
    multas = Multa.objects.select_related('veiculo', 'motorista', 'infracao')

    for multa in multas:
        row = [
            multa.id,
            multa.veiculo.placa if multa.veiculo else '',
            multa.setor or '',
            multa.setor_descricao or '',
            multa.motorista.nome if multa.motorista else '',
            multa.infracao.descricao if multa.infracao else '',
            multa.valor if multa.infracao else '',
            multa.gravidade if multa.infracao else '',
            multa.data_hora_infracao.strftime('%d/%m/%Y %H:%M'),
            multa.local,
            multa.orgao_autuador,
            multa.get_status_multa_display(),
            multa.get_status_pagamento_display(),
            multa.prazo_pagamento.strftime('%d/%m/%Y'),
            multa.data_registro.strftime('%d/%m/%Y %H:%M'),
            '', '', ''
        ]
        ws.append(row)
        current_row = ws.max_row
        domain = request.build_absolute_uri('/')[:-1]
        links = [
            multa.documento_recebido.url if multa.documento_recebido else '',
            multa.comprovante_pagamento.url if multa.comprovante_pagamento else '',
            multa.notificacao_infracao.url if multa.notificacao_infracao else '',
        ]
        for i, file_url in enumerate(links):
            if file_url:
                cell = ws.cell(row=current_row, column=16 + i)
                full_url = f"{domain}{file_url}"
                cell.value = "📎 Abrir Documento"
                cell.hyperlink = full_url
                cell.font = link_font

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=multas.xlsx'
    wb.save(response)
    return response

@user_passes_test(grupo_administrador, login_url='acesso_negado')
def listar_relatorios(request):
    """
    Página para listar relatórios disponíveis.
    """
    return render(request, 'controle/listar_relatorios.html')

@login_required
def criar_conta_pagamento(request):
    if request.method == "POST":
        form = ContaPagamentoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Conta cadastrada com sucesso!")
            return redirect("listar_multas")
    return redirect("listar_multas")

# FIM DO ARQUIVO